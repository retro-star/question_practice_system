from django.shortcuts import render, HttpResponseRedirect, redirect
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
from myapp import models
from django import forms
from openpyxl import load_workbook


class UserModelForm(forms.ModelForm):
    class Meta:
        model = models.User
        fields = ['user_name', 'user_password', 'user_type']

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        for name, filed in self.fields.items():
            filed.widget.attrs = {"placeholder": filed.label}


def register(request):
    if request.method == 'GET':
        return render(request, 'register.html')
    elif request.method == 'POST':
        username = request.POST.get('user_name')
        password = request.POST.get('user_password')
        user = models.User.objects.filter(user_name=username)

        if request.POST.get('login'):
            return redirect('/login/')

        if user:
            error_msg = '该用户已存在'
        elif username:
            user = models.User(user_name=username, user_password=password)
            user.save()
            error_msg = '注册成功,请返回登陆'
        else:
            error_msg = '用户名不能为空'
        return render(request, 'register.html', {'error_msg': error_msg})


def login(request):
    if request.method == 'POST':
        username = request.POST.get('user_name')
        password = request.POST.get('user_password')
        user = models.User.objects.filter(user_name=username)

        if request.POST.get('registered'):
            return redirect('/register/')

        if user:
            if user.filter(user_password=password):
                request.session["info"] = {"id": user[0].id, "username": user[0].user_name,
                                           "usertype": user[0].user_type}
                if user.filter(user_type=1):
                    return redirect('/admin/')
                elif user.filter(user_type=2):
                    return redirect('/teacher/')
                elif user.filter(user_type=3):
                    return redirect('/student/')
            else:
                error_msg = '密码错误'
        else:
            error_msg = '用户名不存在'
        return render(request, 'login.html', {'error_msg': error_msg})
    elif request.method == 'GET':
        return render(request, 'login.html')


def logout(request):
    request.session.clear()
    return redirect("/login/")


def begin(request):
    return render(request, 'login.html')


def admin(request):
    if request.method == 'GET':
        user_list = models.User.objects.all()
        return render(request, 'admin.html', {'user_list': user_list})


def admin_add(request):
    if request.method == 'GET':
        user = UserModelForm()
        return render(request, 'admin_add.html', {'form': user})
    elif request.method == "POST":
        user = UserModelForm(data=request.POST)
        file = request.FILES.get('exc')
        if user.is_valid():
            user.save()
            return redirect('/admin/add/')
        elif file:
            wb = load_workbook(file)
            sheet = wb.worksheets[0]
            for row in sheet.iter_rows(min_row=2):
                username = row[0].value
                userpass = row[1].value
                models.User.objects.create(user_name=username, user_password=userpass)
            return redirect('/admin/add/')
        else:
            return render(request, 'admin_add.html', {"form": user})


def admin_edit(request, nid):
    if request.method == 'GET':
        user = models.User.objects.filter(id=nid).first()
        form = UserModelForm(instance=user)
        return render(request, 'admin_edit.html', {'form': form})

    elif request.method == 'POST':
        user = models.User.objects.filter(id=nid).first()
        form = UserModelForm(data=request.POST, instance=user)
        if form.is_valid():
            form.save()
            return redirect('/admin/')
        else:
            return render(request, 'admin_edit.html', {'form': form})


def admin_delete(request):
    nid = request.GET.get('nid')
    models.User.objects.filter(id=nid).delete()
    return redirect('/admin/')


def admin_course(request):
    if request.method == "GET":
        course = models.Teacher.objects.filter(course_pass=1)
        return render(request, 'admin_course.html', {"course": course})
    elif request.method == "POST":
        cou = models.Teacher.objects.filter(the_course=request.GET.get("course")).first()
        cou.course_pass = request.GET.get("course_pass")
        cou.save()
        return redirect('/admin/course/')


def teacher(request):
    if request.method == "GET":
        info = request.session['info']
        teacher_list = models.Teacher.objects.filter(teacher_name=info["username"])
        return render(request, 'teacher.html', {"teacher_list": teacher_list})


def teacher_delete(request):
    if request.method == "GET":
        nid = request.GET.get("nid")
        models.Teacher.objects.filter(id=nid).delete()
        return redirect('/teacher/')


def teacher_add(request):
    if request.method == "GET":
        return render(request, 'teacher_add.html')
    elif request.method == "POST":
        course_name = request.POST.get("course")
        form = models.Teacher.objects.filter(the_course=course_name)
        if form:
            err_msg = "该课程已存在"
            return render(request, "teacher_add.html", {"err_msg": err_msg})
        else:
            info = request.session["info"]
            teacher_name = info["username"]
            form = models.Teacher(the_course=course_name, teacher_name=models.User.objects.get(
                user_name=teacher_name
            ))
            form.save()
            return redirect("/teacher/")


class ChoiceModeForm(forms.ModelForm):
    class Meta:
        model = models.ChoiceQuestion
        fields = ['question', 'optionA', 'optionB', 'optionC', 'optionD', 'answer', 'its_course']

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        for name, filed in self.fields.items():
            filed.widget.attrs = {"placeholder": filed.label}


class TFModeForm(forms.ModelForm):
    class Meta:
        model = models.TFQuestion
        fields = '__all__'

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        for name, filed in self.fields.items():
            filed.widget.attrs = {"placeholder": filed.label}


class CompletionModeForm(forms.ModelForm):
    class Meta:
        model = models.Completion
        fields = '__all__'

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        for name, filed in self.fields.items():
            filed.widget.attrs = {"placeholder": filed.label}


def question(request):
    if request.method == "GET":
        course = request.GET.get("course")
        choice = models.ChoiceQuestion.objects.filter(its_course=course)
        tf = models.TFQuestion.objects.filter(its_course=course)
        completion = models.Completion.objects.filter(its_course=course)
        cho_form = ChoiceModeForm()
        tf_form = TFModeForm()
        com_form = CompletionModeForm()
        return render(request, "teacher_question.html",
                      {'choice': choice, 'tf': tf, 'completion': completion, 'cho_form': cho_form, 'tf_form': tf_form,
                       'com_form': com_form})


@csrf_exempt
def choice_add(request):
    form = ChoiceModeForm(data=request.POST)
    if form.is_valid():
        form.save()
        return JsonResponse({'status': True})
    return JsonResponse({'status': False, 'error': form.errors})


def choice_delete(request):
    nid = request.GET.get('nid')
    models.ChoiceQuestion.objects.filter(id=nid).delete()
    return


@csrf_exempt
def tf_add(request):
    form = TFModeForm(data=request.POST)
    if form.is_valid():
        form.save()
        return JsonResponse({'status': True})
    print('收到')
    return JsonResponse({'status': False, 'error': form.errors})


def tf_delete(request):
    nid = request.GET.get('nid')
    models.TFQuestion.objects.filter(id=nid).delete()
    return


@csrf_exempt
def com_add(request):
    form = CompletionModeForm(data=request.POST)
    if form.is_valid():
        form.save()
        return JsonResponse({'status': True})
    return JsonResponse({'status': False, 'error': form.errors})


def com_delete(request):
    nid = request.GET.get('nid')
    models.Completion.objects.filter(id=nid).delete()
    return


def student(request):
    if request.method == "GET":
        info = request.session['info']
        stu = models.Student.objects.filter(student_name=info['username'])
        return render(request, 'student.html', {'stu': stu})


def student_add(request):
    if request.method == "GET":
        nid = request.GET.get("nid")
        if nid:
            course = models.Teacher.objects.filter(id=nid).first()
            if models.Student.objects.filter(course=course.the_course):
                err_msg = '该课程已选择'
                course = models.Teacher.objects.filter(course_pass=2)
                return render(request, 'student_add.html', {'course': course, 'err_msg': err_msg})
            else:
                info = request.session['info']
                models.Student.objects.create(course=models.Teacher.objects.get(the_course=course.the_course),
                                              student_name=models.User.objects.get(user_name=info['username']),
                                              teacher_name=models.User.objects.get(user_name=course.teacher_name))
                course = models.Teacher.objects.filter(course_pass=2)
                return render(request, 'student_add.html', {'course': course})
        else:
            course = models.Teacher.objects.filter(course_pass=2)
            return render(request, 'student_add.html', {'course': course})


def test(request):
    if request.method == "GET":
        course = request.GET.get("course")
        choice = models.ChoiceQuestion.objects.filter(its_course=course)
        tf = models.TFQuestion.objects.filter(its_course=course)
        completion = models.Completion.objects.filter(its_course=course)
        return render(request, "test.html", {'choice': choice, 'tf': tf, 'completion': completion, 'course': course})


@csrf_exempt
def cho_anw(request):
    if request.method == 'POST':
        info = request.session['info']
        for key, value in request.POST.items():
            form = models.ChoiceQuestion.objects.filter(question=key).first()
            if form:
                if value == form.answer:
                    tof = 'True'
                else:
                    tof = 'False'
                anw = models.Anw.objects.filter(question=key, student_name=info['username']).first()
                if anw:
                    anw.stu_anw = value
                    anw.tof = tof
                    anw.save()
                else:
                    obj = models.Anw(course=form.its_course,
                                     student_name=models.User.objects.get(user_name=info['username']),
                                     answer=form.answer, stu_anw=value,
                                     question=key, tof=tof, type_of_questions='选择题')
                    obj.save()
    return JsonResponse({'status': True})


@csrf_exempt
def tf_anw(request):
    if request.method == 'POST':
        info = request.session['info']
        for key, value in request.POST.items():
            form = models.TFQuestion.objects.filter(question=key).first()
            if form:
                if bool(value) == form.answer:
                    tof = 'True'
                else:
                    tof = 'False'
                anw = models.Anw.objects.filter(question=key, student_name=info['username']).first()
                if anw:
                    anw.stu_anw = value
                    anw.tof = tof
                    anw.save()
                else:
                    obj = models.Anw(course=form.its_course,
                                     student_name=models.User.objects.get(user_name=info['username']),
                                     answer=form.answer, stu_anw=value,
                                     question=key, tof=tof, type_of_questions='判断题')
                    obj.save()
    return JsonResponse({'status': True})


@csrf_exempt
def com_anw(request):
    if request.method == 'POST':
        info = request.session['info']
        for key, value in request.POST.items():
            form = models.Completion.objects.filter(question=key).first()
            if form:
                if value == form.answer:
                    tof = 'True'
                else:
                    tof = 'False'
                anw = models.Anw.objects.filter(question=key, student_name=info['username']).first()
                if anw:
                    anw.stu_anw = value
                    anw.tof = tof
                    anw.save()
                else:
                    obj = models.Anw(course=form.its_course,
                                     student_name=models.User.objects.get(user_name=info['username']),
                                     answer=form.answer, stu_anw=value,
                                     question=key, tof=tof, type_of_questions='填空题')
                    obj.save()
    return JsonResponse({'status': True})


def grade(request):
    course = request.GET.get('course')
    return render(request, 'student_grade.html', {'course': course})


def grade_all(request):
    course = request.GET.get('course')
    true = 0
    false = 0
    info = request.session['info']
    form = models.Anw.objects.filter(course=course, student_name=info['username'])
    if form:
        for obj in form:
            if obj.tof:
                true = true + 1
            else:
                false = false + 1
        data = [{'value': true, 'name': '正确'}, {'value': false, 'name': '错误'}]
        return JsonResponse({'status': True, 'data': data})


def grade_choice(request):
    course = request.GET.get('course')
    true = 0
    false = 0
    info = request.session['info']
    form = models.Anw.objects.filter(course=course, student_name=info['username'], type_of_questions='选择题')
    if form:
        for obj in form:
            if obj.tof:
                true = true + 1
            else:
                false = false + 1
        data = [{'value': true, 'name': '正确'}, {'value': false, 'name': '错误'}]
        return JsonResponse({'status': True, 'data': data})


def grade_tf(request):
    course = request.GET.get('course')
    true = 0
    false = 0
    info = request.session['info']
    form = models.Anw.objects.filter(course=course, student_name=info['username'], type_of_questions='判断题')
    if form:
        for obj in form:
            if obj.tof:
                true = true + 1
            else:
                false = false + 1
        data = [{'value': true, 'name': '正确'}, {'value': false, 'name': '错误'}]
        return JsonResponse({'status': True, 'data': data})


def grade_com(request):
    course = request.GET.get('course')
    true = 0
    false = 0
    info = request.session['info']
    form = models.Anw.objects.filter(course=course, student_name=info['username'], type_of_questions='填空题')
    if form:
        for obj in form:
            if obj.tof:
                true = true + 1
            else:
                false = false + 1
        data = [{'value': true, 'name': '正确'}, {'value': false, 'name': '错误'}]
        return JsonResponse({'status': True, 'data': data})
